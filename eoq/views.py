from django.shortcuts import render
from .forms import UploadFileForm
import openpyxl, pdb, math

def upload_file(request):
	if request.method == 'POST':
		form = UploadFileForm(request.POST, request.FILES)
		reorder = request.POST.get('reorder')
		holding_cost = request.POST.get('holding_cost')
		wb = openpyxl.load_workbook(request.FILES['myfile'])
		w1 = wb.active	
		max_row = w1.max_row
		cells = []
		total = 0
		average = 0
		count = 0 
		for j in range(2, max_row+1):
			cell = (w1.cell(row=j, column=2)).value
			cells.append(cell)
			total += int(cell)
			count += 1
		average = round(total/count, 2)
		eoq = round(math.sqrt((2*average*float(reorder))/float(holding_cost)))
		context = {'cells':cells, 'average':average, 'reorder':reorder, 'holding_cost':holding_cost, 'eoq':eoq}
		return render(request, 'eoq/total.html', context)
	else:
		form = UploadFileForm()
		return render(request, 'eoq/excel.html', {'form':form})